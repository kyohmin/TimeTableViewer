# ===== Import =====
# GUI Class
import os, csv # For reading data
import tkinter as tk # For GUI Feature
import ttkbootstrap as ttk # For modern tkinter GUI
from ttkbootstrap.dialogs import Messagebox as mb # For error message
from tkinter import filedialog as fd # For file directory

# Data Handler Class and Display Handler Class
import math # For hashtable size
import datetime # For date and time in the schedule

# Export Handler Class
import random # For random color generator
import colorsys # For erandom color generator
from fpdf import FPDF # For editting pdf
from openpyxl import Workbook # For editting Excel
from openpyxl.utils import get_column_letter # # For changing index number to column letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill # For editting components of Excel


"""
Module Download Requirements:
    1. tkinter
    2. ttkbootstrap
    3. openpyxl
    4. fpdf
"""



# ===== Custom Data Structures =====
class Node:
    def __init__(self, data):
        self.data = data
        self.next = None

    def __iter__(self):
        currentNode = self
        while currentNode:
            yield currentNode
            currentNode = currentNode.next

class LinkedList:
    def __init__(self):
        self.head = None # Starting point of the linked list
        self.tail = None # Tail is added for O(1) for appending
    
    def __iter__(self):
        currentNode = self.head
        while currentNode:
            yield currentNode
            currentNode = currentNode.next

    def __contains__(self, node):
        currentNode = self.head
        while currentNode:
            if currentNode.data == node:
                return True
            currentNode = currentNode.next

    def append(self, data):
        newNode = Node(data)
        if self.head == None:
            self.head = newNode
            self.tail = newNode
        
        else:
            self.tail.next = newNode
            self.tail= newNode

    def pop(self):
        if self.head != None:
            tmp = self.head
            self.head = self.head.next
            if self.head == None: self.tail = None
            return tmp.data
        else:
            return None

class HashNode: # Used for Hash table
    def __init__(self, key = -1):
        self.key = key
        self.value = LinkedList() # Value is set to linked list to save schedule classes with same key
        self.next = None # Used when hash value is same, but key is different

class HashTable: # Used for storing schedules with same value.
    def __init__(self, size):
        self.cnt = int(math.floor(size * 1.3)) # Set the list size to reduce collisions
        self.table = [HashNode() for _ in range(self.cnt)]

    def customHash(self, key) -> int: # Generates hashed int value
        key = str(key)
        hashedKey = sum(ord(char) for char in key) % self.cnt
        return hashedKey
    
    def add(self, key, data): # If key exits, append it to linked list. If not, create a Hashnode
        currentNode = self.table[self.customHash(key)]
        while currentNode.next:
            if currentNode.next.key == key:
                currentNode.next.value.append(data)
                return
            currentNode = currentNode.next

        currentNode.next = HashNode(key)
        currentNode.next.value.append(data)

    def get(self, key) -> LinkedList: # Schedule classes are returned in linked list
        currentNode = self.table[self.customHash(key)].next
        while currentNode:
            if currentNode.key == key:
                return currentNode.value
            currentNode = currentNode.next
        return -1

class Stack:
    def __init__(self):
        self.head = None
    
    def __iter__(self):
        currentNode = self.head
        while currentNode:
            yield currentNode
            currentNode = currentNode.next

    def __contains__(self, node):
        currentNode = self.head
        while currentNode:
            if currentNode.data == node:
                return True
            currentNode = currentNode.next

    def push(self, data):
        newNode = Node(data)
        if self.head == None:
            self.head = newNode
        
        else:
            newNode.next = self.head
            self.head = newNode

    def isEmpty(self) -> bool:
        if self.head == None:
            return 1
        else:
            return 0

    def pop(self) -> Node:
        if self.isEmpty():
            return None
        else:
            filteredLL = self.head
            self.head = self.head.next
            filteredLL.next = None
            return filteredLL.data



# ===== Schedule Class =====
class Schedule():
    def __init__(self, module:str, moduleCode:str, cohort:str, course:str, fullPart:str,
                session:str, activityDate:str, scheduledDay:str,startTime:str,endTime:str,
                duration:str,location:str,size:str,lecturer:str,zone:str):
        self.__module = module
        self.__moduleCode = moduleCode
        self.__cohort = cohort
        self.__course = course
        self.__fullPart = fullPart
        self.__session = session
        self.__activityDate = activityDate
        self.__scheduledDay = scheduledDay
        self.__startTime = startTime
        self.__endTime = endTime
        self.__duration = duration
        self.__location = location
        self.__size = size
        self.__lecturer = lecturer
        self.__zone = zone
    
    def get(self, key): # Used for specific value
        value = {
            "module" : self.__module, "moduleCode" : self.__moduleCode, "cohort" : self.__cohort,
            "course" : self.__course, "fullPart" : self.__fullPart, "session" : self.__session,
            "activityDate" : self.__activityDate, "scheduledDay" : self.__scheduledDay,
            "startTime" : self.__startTime, "endTime" : self.__endTime, "duration" : self.__duration,
            "location" : self.__location, "size" : self.__size, "lecturer" : self.__lecturer, "zone" : self.__zone
        }[key]
        
        return value
    
    def getAll(self) -> list: # Used for all values
        li = [
            self.__module, self.__moduleCode, self.__cohort, self.__course, self.__fullPart,
            self.__session, self.__activityDate, self.__scheduledDay, self.__startTime,
            self.__endTime, self.__duration, self.__location, self.__size, self.__lecturer, self.__zone
        ]

        return  li



# ===== Handler Classes =====
class DataHandler:
    def __init__(self):
        self.__schedules = LinkedList()
        self.__filesPathList = []
        self.__rangeList = [] # For initializing hash table. Contains below values in int

        # To hand over the range number to DisplayHandler for intialization of Hash Table
        self.__moduleRange = set()
        self.__moduleCodeRange = set()
        self.__cohortRange = set()
        self.__courseRange = set()
        self.__fullPartRange = set()
        self.__sessionRange = set()
        self.__activityDateRange = set()
        self.__scheduledDayRange = set()
        self.__startTimeRange = set()
        self.__endTimeRange = set()
        self.__durationRange = set()
        self.__locationRange = set()
        self.__sizeRange = set()
        self.__lecturerRange = set()
        self.__zoneRange = set()

    def setSchedules(self):
        self.__schedules.head = None  # Reset the head before storing the schedules
        self.__resetRange() # Reset the range before storing the schedules

        # Read CSV files from the path
        for fileDirectory in self.__filesPathList:
            with open(fileDirectory, "r") as csvfile:
                csvData = csv.reader(csvfile, delimiter=",")
                for row in csvData:
                    if row[0] == '':
                        continue

                    # Extract certain values from the raw data
                    module, moduleCode, cohort, course, fullPart, session = self.__extractData(row[1], row[2])

                    # Set range
                    self.__setRange(
                        module, moduleCode, cohort, course, fullPart, session, self.__dateInput(row[3]),
                        self.__dateInput(row[3]).isoweekday(), self.__timeInput(row[5]),
                        self.__timeInput(row[6]),row[7].lstrip("0"), row[8], int(row[9]), row[10], row[11]
                    )
                    
                    # Store schedule class into linked list
                    self.__schedules.append(
                        Schedule(
                            module, moduleCode, cohort, course, fullPart, session, self.__dateInput(row[3]),
                            self.__dateInput(row[3]).isoweekday(), self.__timeInput(row[5]), self.__timeInput(row[6]),
                            row[7].lstrip("0"), row[8], int(row[9]), row[10], row[11]
                        )
                    )

        # Storing size of range for optimal hashtable size
        self.__rangeList = [len(self.__moduleRange), len(self.__moduleCodeRange), len(self.__cohortRange), len(self.__courseRange), len(self.__fullPartRange), 
                            len(self.__sessionRange), len(self.__activityDateRange), len(self.__scheduledDayRange), len(self.__startTimeRange), len(self.__endTimeRange), 
                            len(self.__durationRange), len(self.__locationRange), len(self.__sizeRange), len(self.__lecturerRange), len(self.__zoneRange)]

    def getSchedules(self) -> Node:
        return self.__schedules.head
    
    def getRangeList(self) -> list:
        return self.__rangeList

    def setFilesPathList(self, filesList:list):
        self.__filesPathList = filesList

    def getFilesPathList(self) -> list:
        return self.__filesPathList

    def setRange(self, schedules):
        # Used externally for finding the range of each values
        self.__resetRange()
        try:
            for i in schedules:
                module, moduleCode, cohort, course, fullPart, session, activityDate, scheduleDay, startTime, endTime, duration, location, size, lecturer, zone = i.data.getAll()
                self.__setRange(module, moduleCode, cohort, course, fullPart, session, activityDate, scheduleDay, startTime, endTime, duration, location, size, lecturer, zone)
        except:
            pass

    def getRange(self, category) -> set: # Used for available options
        value = {
        "module": self.__moduleRange,
        "moduleCode": self.__moduleCodeRange,
        "cohort": self.__cohortRange,
        "course": self.__courseRange,
        "fullPart": self.__fullPartRange,
        "session": self.__sessionRange,
        "activityDate": self.__activityDateRange,
        "scheduledDay": self.__scheduledDayRange,
        "startTime": self.__startTimeRange,
        "endTime": self.__endTimeRange,
        "duration": self.__durationRange,
        "location": self.__locationRange,
        "size": self.__sizeRange,
        "lecturer": self.__lecturerRange,
        "zone": self.__zoneRange,
        }[category]

        return value

    def __dateInput(self, strVal):
        day, month, year = strVal.split("/")
        return datetime.date(int(year),int(month),int(day))

    def __timeInput(self, strVal):
        hour, minute,second = strVal.split(":")
        return datetime.time(int(hour),int(minute),int(second))

    def __setRange(self, module:str, moduleCode:str, cohort:str, course:str, fullPart:str,
                   session:str, activityDate:str, scheduleDay:str, startTime:str, endTime:str,
                   duration:str, location:str, size:str, lecturer:str, zone:str):
        # Internally used for initialization of hash table
        self.__moduleRange.add(module)
        self.__moduleCodeRange.add(moduleCode)
        self.__cohortRange.add(cohort)
        self.__courseRange.add(course)
        self.__fullPartRange.add(fullPart)
        self.__sessionRange.add(session)
        self.__activityDateRange.add(activityDate)
        self.__scheduledDayRange.add(scheduleDay)
        self.__startTimeRange.add(startTime)
        self.__endTimeRange.add(endTime)
        self.__durationRange.add(duration)
        self.__locationRange.add(location)
        self.__sizeRange.add(size)
        self.__lecturerRange.add(lecturer)
        self.__zoneRange.add(zone)

    def __extractData(self,name:str, description:str) -> list:
        course,cohort,fullPart,moduleCode,session = name.split("_")
        cohort = fullPart + "_" + cohort
        module = description[4:].split(" (")[0]
        
        return [module, moduleCode, cohort, course, fullPart, session]
    
    def __resetRange(self):
        self.__moduleRange.clear()
        self.__moduleCodeRange.clear()
        self.__cohortRange.clear()
        self.__courseRange.clear()
        self.__fullPartRange.clear()
        self.__sessionRange.clear()
        self.__activityDateRange.clear()
        self.__scheduledDayRange.clear()
        self.__startTimeRange.clear()
        self.__endTimeRange.clear()
        self.__durationRange.clear()
        self.__locationRange.clear()
        self.__sizeRange.clear()
        self.__lecturerRange.clear()
        self.__zoneRange.clear() 



class DisplayHandler:
    def __init__(self, schedulesHead:Node, rangeList:list):
        # Only the head of the schedule is added for higher performance
        self.__sortedSchedules = schedulesHead

        # Used for filtering data.
        self.__moduleHT = HashTable(rangeList[0])
        self.__moduleCodeHT = HashTable(rangeList[1])
        self.__cohortHT = HashTable(rangeList[2])
        self.__courseHT = HashTable(rangeList[3])
        self.__fullPartHT = HashTable(rangeList[4])
        self.__sessionHT = HashTable(rangeList[5])
        self.__activityDateHT = HashTable(rangeList[6])
        self.__scheduledDayHT = HashTable(rangeList[7])
        self.__startTimeHT = HashTable(rangeList[8])
        self.__endTimeHT = HashTable(rangeList[9])
        self.__durationHT = HashTable(rangeList[10])
        self.__locationHT = HashTable(rangeList[11])
        self.__sizeHT = HashTable(rangeList[12])
        self.__lecturerHT = HashTable(rangeList[13])
        self.__zoneHT = HashTable(rangeList[14])

        # Initialize filters
        if self.__sortedSchedules is not None:
            for schedule in self.__sortedSchedules:
                self.__moduleHT.add(schedule.data.get("module"), schedule.data)
                self.__moduleCodeHT.add(schedule.data.get("moduleCode"), schedule.data)
                self.__cohortHT.add(schedule.data.get("cohort"), schedule.data)
                self.__courseHT.add(schedule.data.get("course"), schedule.data)
                self.__fullPartHT.add(schedule.data.get("fullPart"), schedule.data)
                self.__sessionHT.add(schedule.data.get("session"), schedule.data)
                self.__activityDateHT.add(schedule.data.get("activityDate"), schedule.data)
                self.__scheduledDayHT.add(schedule.data.get("scheduledDay"), schedule.data)
                self.__startTimeHT.add(schedule.data.get("startTime"), schedule.data)
                self.__endTimeHT.add(schedule.data.get("endTime"), schedule.data)
                self.__durationHT.add(schedule.data.get("duration"), schedule.data)
                self.__locationHT.add(schedule.data.get("location"), schedule.data)
                self.__sizeHT.add(schedule.data.get("size"), schedule.data)
                self.__lecturerHT.add(schedule.data.get("lecturer"), schedule.data)
                self.__zoneHT.add(schedule.data.get("zone"), schedule.data)
    
    def getFilteredSchedule(self, category:str, specificValue:str) -> Node:
        filteredLL = None

        if category == "module":
            filteredLL = self.__moduleHT.get(specificValue)
        elif category == "moduleCode":
            filteredLL = self.__moduleCodeHT.get(specificValue)
        elif category == "cohort":
            filteredLL = self.__cohortHT.get(specificValue)
        elif category == "course":
            filteredLL = self.__cohortHT.get(specificValue)
        elif category == "fullPart":
            filteredLL = self.__fullPartHT.get(specificValue)
        elif category == "session":
            filteredLL = self.__sessionHT.get(specificValue)
        elif category == "activityDate":
            filteredLL = self.__activityDateHT.get(specificValue)
        elif category == "scheduledDay":
            filteredLL = self.__scheduledDayHT.get(specificValue)
        elif category == "startTime":
            filteredLL = self.__startTimeHT.get(specificValue)
        elif category == "endTime":
            filteredLL = self.__endTimeHT.get(specificValue)
        elif category == "duration":
            filteredLL = self.__durationHT.get(specificValue)
        elif category == "location":
            filteredLL = self.__locationHT.get(specificValue)
        elif category == "size":
            filteredLL = self.__sizeHT.get(specificValue)
        elif category == "lecturer":
            filteredLL = self.__lecturerHT.get(specificValue)
        elif category == "zone":
            filteredLL = self.__zoneHT.get(specificValue)

        if filteredLL is not None:
            return filteredLL.head

    def rangeSearch(self,start, end, category:str):
        # Finds the schedules between two values and store them
        if category == "time": # For Time - O(n)
            self.sort("startTime")
            head = None
            endExit = False
            for i in self.__sortedSchedules:
                if i.data.get("startTime") >= start: # First schedule found
                    if i.data.get("startTime") > end:
                        endExit = True
                        break
                    head = i
                    break

            if endExit == False:
                self.setCommonSchedules(head)
                self.sort("endTime")
                try:
                    tail = self.__sortedSchedules
                    for i in self.__sortedSchedules:
                        if i.data.get("endTime") > end: # Last schedule found
                            if tail is not None:
                                tail.next = None
                                if tail == self.__sortedSchedules: self.__sortedSchedules = None
                            break
                        tail = i
                except:
                    pass

        else: # For Date - O(n)
            self.sort(category)
            head = None
            endNode = None
            startFound = False
            for i in self.__sortedSchedules:
                if i.data.get(category) >= start and startFound== False:
                    head = i
                    startFound = True
                if i.data.get(category) > end:  
                    if endNode is not None:
                        endNode.next = None
                    break
                endNode = i
            
            self.__sortedSchedules = head

    def getResult(self) -> Node:
        return self.__sortedSchedules
    
    def sort(self, category:str, isAscending=True):
        head = self.__sortedSchedules

        self.__sortedSchedules = self.__mergeSort(head, category)

        # For Descending Order
        if self.__sortedSchedules is not None:
            if isAscending == False:
                stack = Stack()

                for i in self.__sortedSchedules:
                    stack.push(i)

                newHead = currentNode = stack.pop()

                while not stack.isEmpty():
                    currentNode.next = stack.pop()
                    currentNode = currentNode.next

                currentNode.next = None

                self.__sortedSchedules = newHead

    def setCommonSchedules(self, filteredSchedules:Node) -> Node:
        # Used for multiple filtering and range sort features
        node1 = self.__sortedSchedules
        node2 = filteredSchedules
        resultLinkedList = LinkedList()
        availabilityCheck = set()
        while node2 is not None:
            availabilityCheck.add(node2.data)
            node2 = node2.next

        while node1 is not None:
            if node1.data in availabilityCheck:
                resultLinkedList.append(node1.data)
            node1 = node1.next

        self.__sortedSchedules = resultLinkedList.head

    def exportResult(self):
        # Result specifically for export as it requires "sort by date"
        newLinkedList = LinkedList()
        for i in self.__sortedSchedules:
            newLinkedList.append(i.data)
        return self.__mergeSort(newLinkedList.head,"activityDate")

    def __mergeSort(self, head:Node, category:str) -> Node:
        if head is None or head.next is None:
            return head

        mid = self.__split(head)
        left = head
        right = mid.next
        mid.next = None 

        left = self.__mergeSort(left, category)
        right = self.__mergeSort(right, category)

        sortedHead = self.__merge(left, right, category)

        return sortedHead

    def __split(self, head:Node):
        half = head
        end = head

        while end.next is not None and end.next.next is not None:
            half = half.next
            end = end.next.next

        return half

    def __merge(self, left:Node, right:Node, category:str) -> Node: 
        startNode = Node(None)
        currentNode = startNode

        # Loop through to sort between two linked lists
        while left is not None and right is not None:
            if left.data.get(category) < right.data.get(category):
                currentNode.next = left
                left = left.next
            else:
                currentNode.next = right
                right = right.next

            currentNode = currentNode.next

        # After One side is done sorting, add the rest values
        if left is not None:
            currentNode.next = left
        elif right is not None:
            currentNode.next = right

        return startNode.next  # Because first Node was used to create rest of the nodes



class ExportHandler:
    def __init__(self):
        self.__result = LinkedList()
        self.__path = ""
        self.__groupedModules = {}

    def setResult(self, result:LinkedList, path:str):
        self.__result = result
        self.__path = path

    def exportExcel(self):
        # Reset values for repeated process
        self.__startingRow = 12
        self.__weekCnt = 1
        self.__minDate = datetime.date(3000,1,1)
        self.__maxDate = datetime.date(1000,1,1)
        self.__groupedModules = {}

        # Border width predefine
        self.__thick = Side(border_style="medium")
        self.__thin = Side(border_style="thin",color="808080")
        self.__getHeaderData()

        wb = Workbook()
        ws = wb.active

        # First Row with small width
        ws.column_dimensions['A'].width = 5

        # Default Row width
        col = 2
        while col < 10:
            i = get_column_letter(col)
            ws.column_dimensions[i].width = 30
            col += 1

        # PSB logo
        ws.merge_cells('B4:B7')
        ws['B4'].font = Font(size=76,color="902108",name='Times New Roman', bold=True)
        ws['B4'].alignment = Alignment(horizontal="center",vertical="center")
        ws['B4'] = "PSB"

        # ACADEMY logo
        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 20
        ws.merge_cells('B8:B9')
        ws['B8'].font = Font(size=36,color="7F7F7F",name='Calibri', bold=True)
        ws['B8'].alignment = Alignment(horizontal="center",vertical="center")
        ws['B8'] = "Academy"

        # Grouped Module list header
        ws.merge_cells('D4:E4')
        ws['D4'].alignment = Alignment(horizontal="left",vertical="center")
        ws['D4'] = "Module"
        ws['F4'] = "Cohort"
        ws['G4'] = "Lecturer"
        ws['H4'] = "Full-time / Part-time"
        ws['I4'] = "Date Start - Date End"

        ws['D4'].border = Border(left=self.__thick,top=self.__thick,bottom=self.__thick)
        ws['E4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['F4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['G4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['H4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['I4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thick)

        ws['D4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['F4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['G4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['H4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['I4'].fill = PatternFill("solid", start_color="AACDFF")

        # Grouped Module Data Insert
        i = 5
        for _, value in self.__groupedModules.items():
            module, cohort, lecturer, fullPart = value[0], value[1], value[2], value[3]
            ws.merge_cells('D'+str(i)+':'+"E"+str(i))
            ws["D" + str(i)] = module
            ws["F" + str(i)] = cohort
            ws["G" + str(i)] = lecturer
            ws["H" + str(i)] = fullPart
            ws["I" + str(i)] = str(value[5]) + " - " + str(value[6])
            ws["D" + str(i)].fill = PatternFill("solid", start_color=value[4])

            ws["D" + str(i)].border = Border(left=self.__thick,bottom=self.__thin)
            ws["E" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["F" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["G" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["H" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["I" + str(i)].border = Border(bottom=self.__thin,right=self.__thick)
            if len(self.__groupedModules.items())-1 == i-5:
                ws["D" + str(i)].border = Border(left=self.__thick,bottom=self.__thick)
                ws["E" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["F" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["G" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["H" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["I" + str(i)].border = Border(bottom=self.__thick,right=self.__thick)
            ws.row_dimensions[i].height = 20
            i+=1

        # Define the Starting row
        self.__startingRow = 12
        if len(self.__groupedModules.items()) > 5:
            self.__startingRow += len(self.__groupedModules.items())-4

        # Build Main Schedule table
        self.currentDate = self.__minDate
        self.weekStartMarked = False
        self.__createExcelHeading(ws)
        self.__createExcelRow(ws)
        moduleRow = self.__startingRow
        createdRow = 1
        stream = 1
        prevDate = self.__minDate - datetime.timedelta(days=1)

        # Insert schedules
        for schedule in self.__result:
            date = schedule.data.get("activityDate")
            # if new heading has to be created
            while date >= self.currentDate:
                date = schedule.data.get("activityDate")
                self.__createExcelHeading(ws)
                self.__createExcelRow(ws)
                moduleRow = self.__startingRow
                createdRow = 1
                stream = 1

            # if the schedule is the first to go in the date
            if date != prevDate:
                stream = 1
                insertingCell = get_column_letter(date.isoweekday()+2) + str(moduleRow - 6)
                ws[insertingCell] = schedule.data.get("moduleCode") + " (" + schedule.data.get("fullPart")+")"
                ws[insertingCell].fill = PatternFill("solid", start_color=self.__groupedModules[f"{schedule.data.get('module')}{schedule.data.get('cohort')}{schedule.data.get('lecturer')}{schedule.data.get('fullPart')}"][4])
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 5)] = str(schedule.data.get("startTime"))[:-3] + " ~ " + str(schedule.data.get("endTime"))[:-3]
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 4)] = self.__characterLimit(schedule.data.get("lecturer"))
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 3)] = schedule.data.get("location")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 2)] = schedule.data.get("size")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 1)] = schedule.data.get("zone")
            
            # If the schedule has to go to second ~ # rows
            else:
                stream += 1
                # If I need to create a new row for new data
                if createdRow < stream:
                    self.__createExcelRow(ws)
                    createdRow += 1
                # Insert schedule
                insertingCell = get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6))
                ws[insertingCell] = schedule.data.get("moduleCode") + " (" + schedule.data.get("fullPart")+")"
                ws[insertingCell].fill = PatternFill("solid", start_color=self.__groupedModules[f"{schedule.data.get('module')}{schedule.data.get('cohort')}{schedule.data.get('lecturer')}{schedule.data.get('fullPart')}"][4])
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+1)] = str(schedule.data.get("startTime"))[:-3] + " ~ " + str(schedule.data.get("endTime"))[:-3]
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+2)] = self.__characterLimit(schedule.data.get("lecturer"))
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+3)] = schedule.data.get("location")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+4)] = schedule.data.get("size")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+5)] = schedule.data.get("zone")
            prevDate = date


        # Remove the frid lines and headers in Excel for clean view
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = False

        wb.save(self.__path+"/Excel_Timetable.xlsx")

    def exportPDF(self):
        self.__weekStartMarked = False
        self.__weekCnt = 1
        self.__minDate = datetime.date(3000,1,1)
        self.__maxDate = datetime.date(1000,1,1)
        self.__groupedModules = {}

        pdf =  FPDF('L','mm',(500,300))
        pdf.add_page()

        # PSB Logo
        pdf.set_font("times","B",30)
        pdf.set_text_color(158,36,36)
        pdf.cell(33,8,"PSB",ln=True,border=0,align="C")

        # Academy Logo
        pdf.set_font("helvetica","B",12)
        pdf.set_text_color(156,156,156)
        pdf.cell(33,5,"Academy",border=0,ln=True,align="C")
        pdf.cell(33,5,"",ln=True)

        # Table header for grouped modules
        pdf.set_font("helvetica", "B", 7)
        pdf.set_text_color(0,0,0)
        pdf.set_fill_color(170, 205, 255)
        pdf.cell(60,5,"Module", border=True, fill=True)
        pdf.cell(20,5,"Cohort", border=True, fill=True)
        pdf.cell(50,5,"Lecturer", border=True, fill=True)
        pdf.cell(30,5,"Full-time/Part-time", border=True, fill=True)
        pdf.cell(30,5,"Date Start - Date End", border=True, fill=True, ln=True)

        # Table for grouped modules
        self.__getHeaderData(False)
        self.__currentDate = self.__minDate
        pdf.set_font("helvetica", "", 7)
        pdf.set_text_color(0,0,0)
        for i, value in self.__groupedModules.items():
            pdf.set_fill_color(value[4][0],value[4][1],value[4][2])
            pdf.cell(60,5,f"{value[0]}",border=True, fill=True)
            pdf.cell(20,5,f"{value[1]}",border=True, fill=True)
            pdf.cell(50,5,f"{value[2]}",border=True, fill=True)
            pdf.cell(30,5,f"{value[3]}",border=True, fill=True)
            pdf.cell(30,5,f"{value[5]} - {value[6]}",border=True, fill=True, ln=True)

        pdf.cell(30,5,"", ln=True)

        # Header for Schedules table
        while self.__currentDate < self.__maxDate:
            self.__createPDFHeading(pdf)
            self.__createPDFRow(pdf)

        pdf.output(self.__path+"/PDF_Timetable.pdf")

    def __getHeaderData(self, isExcel = True): # Used for finding the grouped schedules
        if isExcel == True: colorVar = "hex"
        else: colorVar = "rgb"

        groupingSchedule = {}
        for i in self.__result:
            strVal = f"{i.data.get('module')}{i.data.get('cohort')}{i.data.get('lecturer')}{i.data.get('fullPart')}"
            # If the schedule is first time recording
            if strVal not in groupingSchedule:
                minDate, maxDate = i.data.get("activityDate"), i.data.get("activityDate")
                groupingSchedule[strVal] = [i.data.get('module')+' '+i.data.get('moduleCode'),i.data.get('cohort'),i.data.get('lecturer'),i.data.get('fullPart'),self.__generateRandomColor(colorVar), minDate, maxDate]

            # If it has been recorded
            else:
                minDate = groupingSchedule[strVal][5]
                maxDate = groupingSchedule[strVal][6]
                # Finding the starting and ending date of each group
                if minDate > i.data.get("activityDate"):
                    minDate = i.data.get("activityDate")
                if maxDate < i.data.get("activityDate"):
                    maxDate = i.data.get("activityDate")

                groupingSchedule.update({strVal:[i.data.get('module')+' '+i.data.get('moduleCode'),i.data.get('cohort'),i.data.get('lecturer'),i.data.get('fullPart'),self.__generateRandomColor(colorVar), minDate, maxDate]})

            if i.data.get("activityDate") < self.__minDate: self.__minDate = i.data.get("activityDate")
            if i.data.get("activityDate") > self.__maxDate: self.__maxDate = i.data.get("activityDate")

        self.__groupedModules = groupingSchedule
    
    def __createExcelHeading(self, ws): # Used for creating heading with days and dates
        # 
        headingFont = Font(size=14, bold=True, underline="single")
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i in range(8):
            currentColumnLetter = get_column_letter(i+2)
            currentCell = currentColumnLetter + str(self.__startingRow)
            bottomCell = currentColumnLetter + str(self.__startingRow+1)
            
            # For the Week columns
            if currentColumnLetter == "B":
                ws.merge_cells('B' + str(self.__startingRow)+':'+'B'+ str(self.__startingRow+1))
                ws[currentCell] = "Week" + str(self.__weekCnt)
                ws[currentCell].border = Border(left=self.__thick, top=self.__thick,right=self.__thick)
                ws[bottomCell].border = Border(left=self.__thick, bottom=self.__thick, right=self.__thick)

            # For the Days and date columns
            else:
                currentCell = currentColumnLetter + str(self.__startingRow)
                ws[currentCell] = days[i-1]
                ws[currentCell].border = Border(top=self.__thick,right=self.__thin)
                ws[bottomCell].border = Border(bottom=self.__thick, right=self.__thin)
                if currentColumnLetter == "I":
                    ws[currentCell].border = Border(top=self.__thick,right=self.__thick)
                    ws[bottomCell].border = Border(bottom=self.__thick,right=self.__thick)
                

            # Font and color on all cells
            ws[currentCell].font = headingFont
            ws[currentCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[bottomCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
            ws[bottomCell].alignment = Alignment(horizontal="center",vertical="center")
        
        # Entering Date
        
        for i in range(7):
            currentColumnLetter = get_column_letter(i+3)
            currentCell = currentColumnLetter + str(self.__startingRow+1)
            if self.__weekCnt == 1 and not self.weekStartMarked:
                if self.currentDate.isoweekday() == i+1:
                    ws[currentCell] = self.currentDate
                    self.weekStartMarked = True
                
            # If start has marked
            else:
                ws[currentCell] = self.currentDate

            if self.weekStartMarked == True:
                self.currentDate += datetime.timedelta(days=1)

        self.__weekCnt += 1
        self.__startingRow += 2

    def __createExcelRow(self,ws): # Used for creating empty schedules row frame
        detailList = ["Module", "Time", "Lecturer", "Location", "Size", "Zone"]
        ws.row_dimensions[self.__startingRow].height = 70
        # Left Column border and data setting
        for i in range(6):
            currentCell = "B" + str(i + self.__startingRow)
            ws[currentCell] = detailList[i]
            ws[currentCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[currentCell].font = Font(bold = True, size=14)
            ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
            ws[currentCell].border = Border(bottom=self.__thin,right=self.__thick,left=self.__thick)

        ws["B" + str(5 + self.__startingRow)].border = Border(bottom=self.__thick,right=self.__thick,left=self.__thick)

        # Setting the font and border for the actual values
        for i in range(7):
            for j in range(6):
                currentColumnLetter = get_column_letter(i+3)
                currentCell = currentColumnLetter + str(j+self.__startingRow)
                ws[currentCell].font = Font(size=14)
                ws[currentCell].border = Border(bottom=self.__thin,right=self.__thin)
                ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
                if j == 5:
                    ws[currentCell].border = Border(bottom=self.__thick,right=self.__thin)
                if i == 6:
                    ws[currentCell].border = Border(bottom=self.__thin,right=self.__thick)
                if j == 5 and i == 6:
                    ws[currentCell].border = Border(bottom=self.__thick,right=self.__thick)

        self.__startingRow += 6

    def __createPDFHeading(self, pdf):
        pdf.set_font("helvetica", "B", 7)
        pdf.set_fill_color(170, 205, 255)
        pdf.cell(30,10, f"Week{self.__weekCnt}", border=True, fill=True, align="C")
        pdf.cell(35,5, "Monday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Tuesday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Wednesday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Thursday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Friday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Saturday", border=True, fill=True, align="C")
        pdf.cell(35,5, "Sunday", border=True, fill=True, align="C", ln=True)
        pdf.cell(30,5, "", border=False,fill=False)
        
        # Adding date in the cell
        for i in range(7):
            pdf.set_font("helvetica", "", 7)
            if self.__weekCnt == 1 and not self.__weekStartMarked:
                if self.__currentDate.isoweekday() == i+1 and i == 6:
                    pdf.cell(35,5,f"{self.__currentDate}",fill=True, border=True, ln=True, align="C")
                    self.__weekStartMarked = True
                elif self.__currentDate.isoweekday() == i+1:
                    pdf.cell(35,5,f"{self.__currentDate}",fill=True,border=True, align="C")
                    self.__weekStartMarked = True
                else:
                    pdf.cell(35,5,"",fill=True,border=True)
            else:
                if i == 6:
                    pdf.cell(35,5,f"{self.__currentDate}",fill=True,border=True, align="C", ln=True)
                else:    
                    pdf.cell(35,5,f"{self.__currentDate}",fill=True,border=True, align="C")

            if self.__weekStartMarked == True:
                self.__currentDate += datetime.timedelta(days=1)
        
        self.__weekCnt += 1
        
    def __createPDFRow(self, pdf, commonSchedules = LinkedList(), common = False):
        pdf.set_font("helvetica", "B", 7)
        # Module row schedule input
        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,25,"Module", fill=True, border = True, align = "C")
        currentWeekSchedules = LinkedList()


        # Adding the current schedule

        if common == False:
            for schedule in self.__result:
                if schedule.data.get("activityDate") < self.__currentDate:
                    currentWeekSchedules.append(schedule.data)
                else:
                    break
        else:
            currentWeekSchedules = commonSchedules
            commonSchedules= LinkedList()

        otherData = LinkedList()

        prev = self.__currentDate
        if currentWeekSchedules.head == None:
            for i in range(7):
                pdf.set_fill_color(255, 255, 255)
                if i == 6:
                    pdf.cell(35,25,f"", fill=True, border = True, align = "C", ln=True)
                else:
                    pdf.cell(35,25,f"", fill=True, border = True, align = "C")
        else:
            for i in range(7):
                # For Module
                if currentWeekSchedules.head != None:
                    if prev != currentWeekSchedules.head.data.get("activityDate"):
                        if currentWeekSchedules.head.data.get("activityDate").isoweekday() == i+1:
                            prev = currentWeekSchedules.head.data.get("activityDate")
                            r,g,b = self.__groupedModules[f"{currentWeekSchedules.head.data.get('module')}{currentWeekSchedules.head.data.get('cohort')}{currentWeekSchedules.head.data.get('lecturer')}{currentWeekSchedules.head.data.get('fullPart')}"][4]
                            pdf.set_fill_color(r,g,b)
                            otherData.append(currentWeekSchedules.head.data)
                            if i == 6:
                                pdf.cell(35,25,f"{currentWeekSchedules.pop().get('moduleCode')}", fill=True, border=True, align="C", ln=True)
                            else:
                                pdf.cell(35,25,f"{currentWeekSchedules.pop().get('moduleCode')}", fill=True, border=True, align="C")

                            while currentWeekSchedules.head != None and prev == currentWeekSchedules.head.data.get("activityDate"):
                                commonSchedules.append(currentWeekSchedules.head.data)
                                currentWeekSchedules.pop()
                        else:
                            pdf.cell(35,25,f"", border = True, align = "C")

                elif currentWeekSchedules.head == None:
                    pdf.set_fill_color(255, 255, 255)
                    if i == 6: pdf.cell(35,25,"", border = True, align = "C", ln=True)
                    else: pdf.cell(35,25,"", border = True, align = "C")


        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,5,"Time", fill=True, border = True, align = "C")
        self.__createOtherPDFData(pdf,otherData,"time")
        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,5,"Lecturer", fill=True, border = True, align = "C")
        self.__createOtherPDFData(pdf,otherData,"lecturer")
        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,5,"Location", fill=True, border = True, align = "C")
        self.__createOtherPDFData(pdf,otherData,"location")
        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,5,"Size", fill=True, border = True, align = "C")
        self.__createOtherPDFData(pdf,otherData,"size")
        pdf.set_fill_color(222, 235, 255)
        pdf.cell(30,5,"Zone", fill=True, border = True, align = "C")
        self.__createOtherPDFData(pdf,otherData,"zone")
        
        if commonSchedules.head == None:
            for i in self.__result:
                if i.data.get("activityDate") >= self.__currentDate:
                    self.__result = i
                    break
        else:
            self.__createPDFRow(pdf,commonSchedules,True)

    def __createOtherPDFData(self,pdf,otherData, category):
        prev = self.__currentDate
        copiedData = LinkedList()
        for i in otherData:
            copiedData.append(i.data)

        if copiedData.head == None:
            for i in range(7):
                pdf.set_fill_color(255, 255, 255)
                if i == 6:
                    pdf.cell(35,5,f"", fill=True, border = True, align = "C", ln=True)
                else:    
                    pdf.cell(35,5,f"", fill=True, border = True, align = "C")
        else:
            for i in range(7):
                # For Module
                if copiedData.head != None:
                    if prev != copiedData.head.data.get("activityDate"):
                        if copiedData.head.data.get("activityDate").isoweekday() == i+1:
                            prev = copiedData.head.data.get("activityDate")

                            if i == 6:
                                if category == "time":
                                    pdf.cell(35,5,f"{copiedData.head.data.get('startTime')} ~ {copiedData.head.data.get('endTime')}", border=True, align="C", ln=True)
                                    copiedData.pop()
                                else:
                                    pdf.cell(35,5,f"{self.__characterLimit(str(copiedData.pop().get(category)))}", border=True, align="C", ln=True)
                            else:
                                if category == "time":
                                    pdf.cell(35,5,f"{copiedData.head.data.get('startTime')} ~ {copiedData.head.data.get('endTime')}", border=True, align="C")
                                    copiedData.pop()
                                else:
                                    pdf.cell(35,5,f"{self.__characterLimit(str(copiedData.pop().get(category)))}", border=True, align="C")

                            while copiedData.head != None and prev == copiedData.head.data.get("activityDate"):
                                copiedData.pop()
                        else:
                            # After End
                            if i == 6:
                                pdf.cell(35,5,f"", border = True, align = "C",ln=True)
                            else:
                                pdf.cell(35,5,f"", border = True, align = "C")

                elif copiedData.head == None:
                    if i == 6: pdf.cell(35,5,"", border = True, align = "C", ln=True)
                    else: pdf.cell(35,5,"", border = True, align = "C")

    def __characterLimit(self, strVal) -> str:
        # If the character is too big, only display first and last name
        if len(strVal) > 25:
            return strVal.split()[0] + " "+ strVal.split()[-1]
        else:
            return strVal
        
    def __generateRandomColor(self, resultType="hex"):
        # Generate random color for grouping modules
        h,s,l = random.random()*360, random.random(), 0.9
        r,g,b = [int(256*i) for i in colorsys.hls_to_rgb(h,l,s)]
        
        if resultType == "hex":
            result = '{:02x}{:02x}{:02x}'.format(r, g, b)
        elif resultType == "rgb":
            result = [int(r),int(g),int(b)]
            
        return result



# ===== GUI Class =====
class GUI(ttk.Window):
    def __init__(self):
        # Initialize Data Handler when program initialize
        self.__dataHandler = DataHandler()
        self.__exportHandler = ExportHandler()

        # Make a window setting
        self.__initProgram() # Set all GUI components and features
        self.__currentFrame = self.__mainPage # Used for tracking current page
        self.__showPage(self.__mainPage) # Start the page with main page
        self.mainloop()

    def __initProgram(self):
        super().__init__()
        theme = ttk.Style()
        theme.theme_use("flatly")
        self.resizable(False,False)
        self.title("Timetable Viewer")
        self.geometry("750x500+500+250")
        self.rowconfigure(0, weight = 1)
        self.columnconfigure(0, weight=1)

        # Initializing the pages
        self.__mainPage = ttk.Frame(self)
        self.__loadPage = ttk.Frame(self)
        self.__viewPage = ttk.Frame(self, width = 1500, height=500)

        # Set each pages
        self.__initMainPage()
        self.__initLoadPage()
        self.__initViewPage()

    def __initMainPage(self):
        # Setting Page
        self.__mainPage.grid(row=0, column=0, sticky="nswe")
        self.__mainPage.rowconfigure(0,weight= 5)
        self.__mainPage.rowconfigure(1,weight= 6)
        self.__mainPage.columnconfigure(0,weight=1)
        self.__mainPage.columnconfigure(1,weight=1)
        self.__mainPage.columnconfigure(2,weight=1)
        self.__mainPage.columnconfigure(3,weight=1)
        self.__mainPage.columnconfigure(4,weight=1)

        # Predefining title and button style
        titleFont = ("Arial", 40, "bold")
        myButtonStyle = ttk.Style()
        myButtonStyle.configure('my.TButton', font=("Arial", 20))

        # Specific Setting
        self.__mainPageLabel = ttk.Label(self.__mainPage, text="Timetable Viewer", font = titleFont)
        self.__mainPageSeparator = ttk.Separator(self.__mainPage, bootstyle="primary")
        self.__startButton = ttk.Button(self.__mainPage, text="START", style="my.TButton", command=lambda:[self.__showPage(self.__loadPage)])
        self.__exitBtn = ttk.Button(self.__mainPage, text="EXIT", style="my.TButton", command=self.destroy)

        # Show on main page
        self.__mainPageLabel.grid(row = 0 , column=2, sticky = "s", pady=(0,30))
        self.__mainPageSeparator.grid(row = 0, column=2, sticky= "swe", pady=(0,0), padx=200)
        self.__startButton.grid(row = 1, column = 2, sticky = "nwe", pady=(60), ipady = 10, padx=150)
        self.__exitBtn.grid(row = 1, column = 2, sticky = "nwe", pady=(135,0), ipady=10, padx=150)

    def __initLoadPage(self):
        # Setting Page
        self.__loadPage.grid(row=0, column=0, sticky="nswe")
        self.__loadPage.rowconfigure(0,weight=4)
        self.__loadPage.rowconfigure(1,weight=4)
        self.__loadPage.rowconfigure(2,weight=2)
        self.__loadPage.columnconfigure(0,weight=2)
        self.__loadPage.columnconfigure(1,weight=3)
        self.__loadPage.columnconfigure(2,weight=2)

        # Predefining title and button style
        titleFont = ("Arial", 40, "bold")
        midButtonStyle = ttk.Style()
        deleteButtonStyle = ttk.Style()
        clearButtonStyle = ttk.Style()
        midButtonStyle.configure('mid.TButton', font=("Arial", 15))
        deleteButtonStyle.configure('s.danger.Outline.TButton', font=("Arial", 10), bootstyle="dark-outline")
        clearButtonStyle.configure('s.warning.Outline.TButton', font=("Arial", 10), bootstyle="dark-outline")

        # Specific Setting
        column = ("indexNum", "fileName")
        self.__loadPageLabel = ttk.Label(self.__loadPage, text="Locate The CSV Files", anchor="center", font=titleFont)
        self.__openFolderButton = ttk.Button(self.__loadPage, text="Open Folder", bootstyle="dark-outline", command=lambda:[self.__queryFolderPath()])
        self.__fileTree = ttk.Treeview(self.__loadPage, columns = column, show="headings", bootstyle="dark")
        self.__scrollbar = ttk.Scrollbar(self.__loadPage, orient="vertical",command=self.__fileTree.yview)
        self.__deleteBtn = ttk.Button(self.__loadPage, text="Delete", style="s.danger.Outline.TButton", command=lambda:[self.__deleteFile()])
        self.__clearBtn = ttk.Button(self.__loadPage, text="Clear", style="s.warning.Outline.TButton", command=lambda:[self.__clearAllFiles()])
        self.__backHomeButton = ttk.Button(self.__loadPage, text="Back", style="mid.TButton", command=lambda:[self.__showPage(self.__mainPage)])
        self.__nextButton = ttk.Button(self.__loadPage, text="Next", style="mid.TButton", command=lambda:[self.__showPage(self.__viewPage)])

        # Show on screen    
        self.__loadPageLabel.grid(row=0,column=0, sticky="nwe", pady=(30,0), columnspan=3)
        self.__openFolderButton.grid(row=0,column=1, sticky="s", pady=(0,10), ipady=1, ipadx=40)
        self.__fileTree.grid(row=1,column=1, sticky="ns")
        self.__scrollbar.grid(row=1,column=1,sticky="ens", padx=(0,33))
        self.__fileTree.configure(yscrollcommand=self.__scrollbar.set)
        self.__deleteBtn.grid(row=2, column=1, sticky="n", pady=(5,0), padx=(100,0))
        self.__clearBtn.grid(row=2, column=1, sticky="n", pady=(5,0), padx=(0,100))
        self.__backHomeButton.grid(row=2,column=0, sticky="s", pady=(0,20), padx=(0,0), ipadx=20, ipady=5)
        self.__nextButton.grid(row=2,column=2, sticky="s", pady=(0,20), padx=(0,0), ipadx=20, ipady=5)

        # Setting table for list of files
        self.__fileTree.column("indexNum", anchor="w")
        self.__fileTree.column("fileName", anchor="w")
        self.__fileTree.heading("indexNum", text="#", anchor="w")
        self.__fileTree.heading("fileName", text="File Name", anchor="w")
        self.__fileTree.column("indexNum", width=30, minwidth=30, stretch=tk.NO)
        self.__fileTree.column("fileName", width=300, minwidth=300, stretch=tk.NO)
        
    def __initViewPage(self):
        # Setting Page
        self.__viewPage.grid(row=0, column=0, sticky="nswe")
        self.__viewPage.columnconfigure(0,weight=2)
        self.__viewPage.columnconfigure(1,weight=8)
        self.__viewPage.rowconfigure(0,weight=1)

        # Split the screen in half
        self.__filterFrame = ttk.Frame(self.__viewPage,bootstyle="light")
        self.__tableFrame = ttk.Frame(self.__viewPage)

        # Setting for the splitted screen
        self.__filterFrame.grid(row=0,column=0,sticky="nswe")
        self.__tableFrame.grid(row=0,column=1,sticky="nswe")

        # Setting for the filter 
        self.__filterFrame.columnconfigure(0,weight=1, uniform='filter')
        self.__filterFrame.columnconfigure(1,weight=1, uniform='filter')
        self.__filterFrame.columnconfigure(2,weight=1, uniform='filter')
        self.__filterFrame.rowconfigure(0,weight=1)
        self.__filterFrame.rowconfigure(1,weight=1)
        self.__filterFrame.rowconfigure(2,weight=1)
        self.__filterFrame.rowconfigure(3,weight=1)
        self.__filterFrame.rowconfigure(4,weight=1)
        self.__filterFrame.rowconfigure(5,weight=1)
        self.__filterFrame.rowconfigure(6,weight=1)
        self.__filterFrame.rowconfigure(7,weight=1)
        self.__filterFrame.rowconfigure(8,weight=1)
        self.__filterFrame.rowconfigure(9,weight=1)

        # Setting for the table frame
        self.__tableFrame.columnconfigure(0, weight=1)
        self.__tableFrame.rowconfigure(0, weight=1)
        self.__tableFrame.rowconfigure(1, weight=5)
        

        # Predefining font and button styles
        titleFont = ("Arial", 20, "bold")
        resetButton = ttk.Style()
        menuButtonStyle = ttk.Style()
        menuButtonStyle.configure("new.primary.Outline.TMenubutton", font=("Arial",13, "bold"))
        resetButton.configure("reset.danger.Outline.TButton", font=("Arial",10))

        # Setting table
        column = ["index","module", "moduleCode", "cohort", "course", "fullPart", "session", "activityDate", "scheduledDay", "startTime", "endTime", "duration","location","size","lecturer","zone"]
        self.__scheduleViewer = ttk.Treeview(self.__tableFrame, columns = column, show="headings", bootstyle="primary")
        self.__scheduleViewer.grid(row=1,column=0, padx=(10,21),pady=(0,21), sticky="nswe")
        self.__scheduleViewer.column("index", anchor="w", minwidth=40, width=40)
        self.__scheduleViewer.column("module", anchor="w", minwidth=170, width=10)
        self.__scheduleViewer.column("moduleCode", anchor="w", minwidth=100, width=10)
        self.__scheduleViewer.column("cohort", anchor="w", minwidth=60, width=10)
        self.__scheduleViewer.column("course", anchor="w", minwidth=100, width=10)
        self.__scheduleViewer.column("fullPart", anchor="w", minwidth=70, width=10)
        self.__scheduleViewer.column("session", anchor="w", minwidth=70, width=10)
        self.__scheduleViewer.column("activityDate", anchor="w", minwidth=100, width=10)
        self.__scheduleViewer.column("scheduledDay", anchor="w", minwidth=80, width=10)
        self.__scheduleViewer.column("startTime", anchor="w", minwidth=80, width=10)
        self.__scheduleViewer.column("endTime", anchor="w", minwidth=80, width=10)
        self.__scheduleViewer.column("duration", anchor="w", minwidth=70, width=10)
        self.__scheduleViewer.column("location", anchor="w", minwidth=60, width=10)
        self.__scheduleViewer.column("size", anchor="w", minwidth=50, width=10)
        self.__scheduleViewer.column("lecturer", anchor="w", minwidth=100, width=10)
        self.__scheduleViewer.column("zone", anchor="w", minwidth=50, width=10)

        self.__scheduleViewer.heading("index", text="#", anchor="w")
        self.__scheduleViewer.heading("module", text="Module", anchor="w")
        self.__scheduleViewer.heading("moduleCode", text="Module Code",anchor="w")
        self.__scheduleViewer.heading("cohort", text="Cohort",anchor="w")
        self.__scheduleViewer.heading("course", text="Course",anchor="w")
        self.__scheduleViewer.heading("fullPart", text="Full/Part",anchor="w")
        self.__scheduleViewer.heading("session", text="Session",anchor="w")
        self.__scheduleViewer.heading("activityDate", text="Date",anchor="w")
        self.__scheduleViewer.heading("scheduledDay", text="Day",anchor="w")
        self.__scheduleViewer.heading("startTime", text="Start Time",anchor="w")
        self.__scheduleViewer.heading("endTime", text="End Time",anchor="w")
        self.__scheduleViewer.heading("duration", text="Duration",anchor="w")
        self.__scheduleViewer.heading("location", text="Location",anchor="w")
        self.__scheduleViewer.heading("size", text="Class Size",anchor="w")
        self.__scheduleViewer.heading("lecturer", text="Lecturer",anchor="w")
        self.__scheduleViewer.heading("zone", text="Zone",anchor="w")

        # Setting table scrollbar
        self.__scheduleScrollVerBar = ttk.Scrollbar(self.__tableFrame, orient="vertical",command=self.__scheduleViewer.yview)
        self.__scheduleScrollHorBar = ttk.Scrollbar(self.__tableFrame, orient="horizontal",command=self.__scheduleViewer.xview)
        self.__scheduleScrollVerBar.grid(row=1,column=0,sticky="ens",padx=(0,10),pady=(0,21))
        self.__scheduleScrollHorBar.grid(row=1,column=0,sticky="wes", padx=(10,10),pady=(0,10))
        self.__scheduleViewer.configure(yscrollcommand=self.__scheduleScrollVerBar.set)
        self.__scheduleViewer.configure(xscrollcommand=self.__scheduleScrollHorBar.set)
 
        # Specific Setting
        self.__filterLabel = ttk.Label(self.__filterFrame, text="Filter Schedules", font=titleFont, bootstyle="inverse-light")
        self.__RangeLabel = ttk.Label(self.__filterFrame, text="Range Schedules", font=titleFont, bootstyle="inverse-light")
    
        # Setting components
        self.__startDate = ttk.DateEntry(self.__filterFrame, width=8, dateformat="%d-%m-%Y")
        self.__endDate = ttk.DateEntry(self.__filterFrame, width=8, dateformat="%d-%m-%Y")
        self.__applyDateButton = ttk.Button(self.__filterFrame, text="Apply Date", bootstyle="primary",command=lambda:[self.__queryFilterDateTime("date",self.__startDate.entry.get(),self.__endDate.entry.get())])
        self.__applyTimeButton = ttk.Button(self.__filterFrame, text="Apply Time", bootstyle="primary",command=lambda:[self.__queryFilterDateTime("time",self.__startTimeEntry.get(),self.__endTimeEntry.get())])
        self.__startTimeEntry = ttk.Entry(self.__filterFrame, width=13)
        self.__startTimeEntry.insert(0,"01:00")
        self.__endTimeEntry = ttk.Entry(self.__filterFrame, width=13)
        self.__endTimeEntry.insert(0,"23:00")
        self.__backLoadButton = ttk.Button(self.__filterFrame, text="Back", bootstyle="primary",command=lambda:[self.__showPage(self.__loadPage)])
        self.__scheduleLabel = ttk.Label(self.__tableFrame, text="Schedules", font=titleFont)
        self.__resetSchedulesButton = ttk.Button(self.__filterFrame, text="Reset", style="reset.danger.Outline.TButton",command=lambda:[self.__showPage(self.__viewPage)])

        # Export Menu button
        self.__exportMenu = ttk.Menubutton(self.__filterFrame, style="primary.Outline.TMenubutton", text="Export")
        exportOptions = ttk.Menu(self.__exportMenu)
        exportOptions.add_radiobutton(label="Excel", command=lambda:[self.__exportExcel()])
        exportOptions.add_radiobutton(label="PDF", command=lambda:[self.__exportPDF()])
        self.__exportMenu['menu'] = exportOptions

        # Setting filter options
        self.__moduleMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Module")
        self.__moduleCodeMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Module Code")
        self.__cohortMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Cohort")
        self.__courseMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Course")
        self.__fullPartMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Full / Part")
        self.__sessionMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Session")
        self.__activityDateMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Date")
        self.__scheduledDayMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Day")
        self.__startTimeMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Start Time")
        self.__endTimeMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="End Time")
        self.__durationMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Duration")
        self.__locationMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Location")
        self.__sizeMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Size")
        self.__lecturerMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Lecturer")
        self.__zoneMenu = ttk.Menubutton(self.__filterFrame, style="new.primary.Outline.TMenubutton", text="Zone")

        # Sort By buttons setting
        self.__sortIn = tk.BooleanVar()
        self.__sortBySelection = "Module"
        self.__ascButton = ttk.Radiobutton(self.__tableFrame, variable=self.__sortIn, text="Ascending", value=True, command=lambda:[self.__selectedValue("sortBy",self.__sortBySelection,True)])
        self.__desButton = ttk.Radiobutton(self.__tableFrame, variable=self.__sortIn, text="Descending", value=False, command=lambda:[self.__selectedValue("sortBy",self.__sortBySelection,False)])
        self.__sortByMenu = ttk.Menubutton(self.__tableFrame, style="new.primary.Outline.TMenubutton", text="Sort By")

        # Setting table headings
        self.__sortByVar = tk.StringVar()
        sortByOptions = ttk.Menu(self.__sortByMenu)
        for x in ["Module","Module Code","Cohort","Course","Full/Part","Session","Date","Day","Start Time","End Time","Duration","Location","Size","Lecturer","Zone"]:
            sortByOptions.add_radiobutton(label=x, variable=self.__sortByVar, command=lambda x=x:self.__selectedValue("sortBy",x,self.__sortIn.get()))
        self.__sortByMenu['menu'] = sortByOptions

        # Show on frame
        self.__filterLabel.grid(row=0, column=0,columnspan=3, sticky="w", padx=(10,0))
        self.__RangeLabel.grid(row=6,column=0,columnspan=3,sticky="w",padx=(10,0))
        self.__scheduleLabel.grid(row=0, column=0, sticky="w",padx=(10,0))
        self.__startDate.grid(row=7,column=0, ipadx=10)
        self.__endDate.grid(row=7,column=1, ipadx=10)
        self.__applyDateButton.grid(row=7,column=2)
        self.__startTimeEntry.grid(row=8,column=0)
        self.__endTimeEntry.grid(row=8,column=1)
        self.__applyTimeButton.grid(row=8,column=2)
        self.__backLoadButton.grid(row=9, column=0, sticky="s", pady=(0,10), padx=(0,0), ipadx=20, ipady=5)
        self.__resetSchedulesButton.grid(row=9, column=1, sticky="s", pady=(0,15), padx=40, ipadx=20, ipady=1)
        self.__exportMenu.grid(row=9,column=2,stick="wes", pady=(0,10), padx=30, ipadx=10, ipady=5)
        self.__moduleMenu.grid(row=1,column=0,sticky="we",padx=10)
        self.__moduleCodeMenu.grid(row=1,column=1,sticky="we",padx=10)
        self.__cohortMenu.grid(row=1,column=2,sticky="we",padx=10)
        self.__courseMenu.grid(row=2,column=0,sticky="we",padx=10)
        self.__fullPartMenu.grid(row=2,column=1,sticky="we",padx=10)
        self.__sessionMenu.grid(row=2,column=2,sticky="we",padx=10)
        self.__activityDateMenu.grid(row=3, column=0,sticky="we",padx=10)
        self.__scheduledDayMenu.grid(row=3,column=1,sticky="we",padx=10)
        self.__locationMenu.grid(row=3,column=2,sticky="we",padx=10)
        self.__startTimeMenu.grid(row=4,column=0,sticky="we",padx=10)
        self.__endTimeMenu.grid(row=4,column=1,sticky="we",padx=10)
        self.__durationMenu.grid(row=4,column=2,sticky="we",padx=10)
        self.__lecturerMenu.grid(row=5,column=0,sticky="we",padx=10)
        self.__sizeMenu.grid(row=5,column=1,sticky="we",padx=10)
        self.__zoneMenu.grid(row=5,column=2,sticky="we",padx=10)
        self.__sortByMenu.grid(row=0,column=0,sticky="e",padx=(0,21),pady=(0,0))
        self.__ascButton.grid(row=0,column=0,padx=(0,250), sticky="e")
        self.__desButton.grid(row=0,column=0,padx=(0,150), sticky="e")

    def __selectedValue(self, category:str, value, isASC = True):
        # Receive what filter user selected and sends it to the display handler for data processing
        if category == "sortBy":
            newCategory = {
                "Module" : "module",
                "Module Code" : "moduleCode",
                "Cohort" : "cohort",
                "Course" : "course",
                "Full/Part" : "fullPart",
                "Session" : "session",
                "Date" : "activityDate",
                "Day" : "scheduledDay",
                "Start Time" : "startTime",
                "End Time" : "endTime",
                "Duration" : "duration",
                "Location" : "location",
                "Size" : "size",
                "Lecturer" : "lecturer",
                "Zone" : "zone",
            }[value]
            self.__sortBySelection = value
            self.__displayHandler.sort(newCategory, isASC)
        else:
            self.__displayHandler.setCommonSchedules(self.__displayHandler.getFilteredSchedule(category,value))
            self.__dataHandler.setRange(self.__displayHandler.getResult())
        

        # Post Filtering or Sorting
        self.__displaySchedules()
        self.__updateOptions()

    def __updateOptions(self):
        # Style settings
        disabledButtonStyle = ttk.Style()
        enabledButtonStyle = ttk.Style()
        disabledButtonStyle.configure("dis.dark.TMenubutton", font=("Arial",13, "bold"))
        enabledButtonStyle.configure("en.primary.Outline.TMenubutton", font=("Arial",13, "bold"))

        # Module options renewal
        moduleOptions = ttk.Menu(self.__moduleMenu)
        moduleVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("module")):
            moduleOptions.add_radiobutton(label=x, variable=moduleVar, command=lambda x=x:self.__selectedValue("module",x))
        self.__moduleMenu['menu'] = moduleOptions

        if len(self.__dataHandler.getRange("module")) <= 1:
            self.__moduleMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__moduleMenu.configure(style="en.primary.Outline.TMenubutton")

        # Module Code options renewal
        moduleCodeOptions = ttk.Menu(self.__moduleCodeMenu)
        moduleCodeVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("moduleCode")):
            moduleCodeOptions.add_radiobutton(label=x, variable=moduleCodeVar, command=lambda x=x:self.__selectedValue("moduleCode",x))
        self.__moduleCodeMenu['menu'] = moduleCodeOptions

        if len(self.__dataHandler.getRange("moduleCode")) <= 1:
            self.__moduleCodeMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__moduleCodeMenu.configure(style="en.primary.Outline.TMenubutton")

        # Cohort options renewal
        cohortOptions = ttk.Menu(self.__cohortMenu)
        cohortVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("cohort")):
            cohortOptions.add_radiobutton(label=x, variable=cohortVar, command=lambda x=x:self.__selectedValue("cohort",x))
        self.__cohortMenu['menu'] = cohortOptions

        if len(self.__dataHandler.getRange("cohort")) <= 1:
            self.__cohortMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__cohortMenu.configure(style="en.primary.Outline.TMenubutton")

        # Course options renewal
        courseOptions = ttk.Menu(self.__courseMenu)
        courseVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("course")):
            courseOptions.add_radiobutton(label=x, variable=courseVar, command=lambda x=x:self.__selectedValue("course",x))
        self.__courseMenu['menu'] = courseOptions

        if len(self.__dataHandler.getRange("course")) <= 1:
            self.__courseMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__courseMenu.configure(style="en.primary.Outline.TMenubutton")

        # Full/Part options renewal
        fullPartOptions = ttk.Menu(self.__fullPartMenu)
        fullPartVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("fullPart")):
            fullPartOptions.add_radiobutton(label=x, variable=fullPartVar, command=lambda x=x:self.__selectedValue("fullPart",x))
        self.__fullPartMenu['menu'] = fullPartOptions

        if len(self.__dataHandler.getRange("fullPart")) <= 1:
            self.__fullPartMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__fullPartMenu.configure(style="en.primary.Outline.TMenubutton")

        # Session options renewal
        sessionOptions = ttk.Menu(self.__sessionMenu)
        sessionVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("session")):
            sessionOptions.add_radiobutton(label=x, variable=sessionVar, command=lambda x=x:self.__selectedValue("session",x))
        self.__sessionMenu['menu'] = sessionOptions

        if len(self.__dataHandler.getRange("session")) <= 1:
            self.__sessionMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__sessionMenu.configure(style="en.primary.Outline.TMenubutton")

        # Activity Date options renewal
        activityDateOptions = ttk.Menu(self.__activityDateMenu)
        activityDateVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("activityDate")):
            activityDateOptions.add_radiobutton(label=x, variable=activityDateVar, command=lambda x=x:self.__selectedValue("activityDate",x))
        self.__activityDateMenu['menu'] = activityDateOptions

        if len(self.__dataHandler.getRange("activityDate")) <= 1:
            self.__activityDateMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__activityDateMenu.configure(style="en.primary.Outline.TMenubutton")

        # Scheduled Day options renewal
        day = ["", "Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        scheduledDayOptions = ttk.Menu(self.__scheduledDayMenu)
        scheduledDayVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("scheduledDay")):
            scheduledDayOptions.add_radiobutton(label=day[x], variable=scheduledDayVar, command=lambda x=x:self.__selectedValue("scheduledDay",x))
        self.__scheduledDayMenu['menu'] = scheduledDayOptions

        if len(self.__dataHandler.getRange("scheduledDay")) <= 1:
            self.__scheduledDayMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__scheduledDayMenu.configure(style="en.primary.Outline.TMenubutton")

        # Start Time options renewal
        startTimeOptions = ttk.Menu(self.__startTimeMenu)
        startTimeVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("startTime")):
            startTimeOptions.add_radiobutton(label=x, variable=startTimeVar, command=lambda x=x:self.__selectedValue("startTime",x))
        self.__startTimeMenu['menu'] = startTimeOptions

        if len(self.__dataHandler.getRange("startTime")) <= 1:
            self.__startTimeMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__startTimeMenu.configure(style="en.primary.Outline.TMenubutton")

        # End Time options renewal
        endTimeOptions = ttk.Menu(self.__endTimeMenu)
        endTimeVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("endTime")):
            endTimeOptions.add_radiobutton(label=x, variable=endTimeVar, command=lambda x=x:self.__selectedValue("endTime",x))
        self.__endTimeMenu['menu'] = endTimeOptions

        if len(self.__dataHandler.getRange("endTime")) <= 1:
            self.__endTimeMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__endTimeMenu.configure(style="en.primary.Outline.TMenubutton")

        # Duration options renewal
        durationOptions = ttk.Menu(self.__durationMenu)
        durationVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("duration")):
            durationOptions.add_radiobutton(label=x, variable=durationVar, command=lambda x=x:self.__selectedValue("duration",x))
        self.__durationMenu['menu'] = durationOptions
        
        if len(self.__dataHandler.getRange("duration")) <= 1:
            self.__durationMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__durationMenu.configure(style="en.primary.Outline.TMenubutton")

        # Location options renewal
        locationOptions = ttk.Menu(self.__locationMenu)
        locationVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("location")):
            locationOptions.add_radiobutton(label=x, variable=locationVar, command=lambda x=x:self.__selectedValue("location",x))
        self.__locationMenu['menu'] = locationOptions

        if len(self.__dataHandler.getRange("location")) <= 1:
            self.__locationMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__locationMenu.configure(style="en.primary.Outline.TMenubutton")

        # Size options renewal
        sizeOptions = ttk.Menu(self.__sizeMenu)
        sizeVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("size")):
            sizeOptions.add_radiobutton(label=x, variable=sizeVar, command=lambda x=x:self.__selectedValue("size",x))
        self.__sizeMenu['menu'] = sizeOptions

        if len(self.__dataHandler.getRange("size")) <= 1:
            self.__sizeMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__sizeMenu.configure(style="en.primary.Outline.TMenubutton")

        # Lecturer options renewal
        lecturerOptions = ttk.Menu(self.__lecturerMenu)
        lecturerVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("lecturer")):
            lecturerOptions.add_radiobutton(label=x, variable=lecturerVar, command=lambda x=x:self.__selectedValue("lecturer",x))
        self.__lecturerMenu['menu'] = lecturerOptions

        if len(self.__dataHandler.getRange("lecturer")) <= 1:
            self.__lecturerMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__lecturerMenu.configure(style="en.primary.Outline.TMenubutton")

        # Zone options renewal
        zoneOptions = ttk.Menu(self.__zoneMenu)
        zoneVar = tk.StringVar()
        for x in self.__insertionSort(self.__dataHandler.getRange("zone")):
            zoneOptions.add_radiobutton(label=x, variable=zoneVar, command=lambda x=x:self.__selectedValue("zone",x))
        self.__zoneMenu['menu'] = zoneOptions

        if len(self.__dataHandler.getRange("zone")) <= 1:
            self.__zoneMenu.configure(style="dis.dark.TMenubutton")
        else:
            self.__zoneMenu.configure(style="en.primary.Outline.TMenubutton")

    def __queryFilterDateTime(self, category:str, start, end):
        # Filter schedules between start and end values.
        try:
            if category == "date":
                dateformat = "%d-%m-%Y"
                start = datetime.datetime.strptime(f"{start}",dateformat).date()
                end = datetime.datetime.strptime(f"{end}",dateformat).date()
                if start > end: mb.show_warning("Starting date cannot exceed ending date.", "Wrong date input")
                self.__displayHandler.rangeSearch(start,end,"activityDate")
            elif category == "time":
                dateformat = "%H:%M"
                start = datetime.datetime.strptime(f"{start}",dateformat).time()
                end = datetime.datetime.strptime(f"{end}",dateformat).time()
                if start > end: mb.show_warning("Starting time cannot exceed ending time.", "Wrong time input")
                self.__displayHandler.rangeSearch(start,end,"time")
            self.__dataHandler.setRange(self.__displayHandler.getResult())
            self.__displaySchedules()
            self.__updateOptions()
        except:
            mb.show_error("You entered wrong value\nPlease try again", "Wrong input error")
 
    def __queryFolderPath(self):
        # Receives the filder path and extracts file path. Then send it to the data handler for extracting raw data
        folderPath = fd.askdirectory()
        filesList = []
        if folderPath != "":
            for file in os.listdir(folderPath):
                if file.endswith(".csv"):
                    filesList.append(os.path.join(folderPath, file))
            self.__dataHandler.setFilesPathList(filesList)
            self.__displayFiles()

            if len(filesList) == 0:
                mb.show_error("You chose a folder with no CSV files\nPlease choose a folder with CSV files", "No Files Error")

    def __displayFiles(self):
        for i in self.__fileTree.get_children():
            self.__fileTree.delete(i)

        filesList = self.__dataHandler.getFilesPathList()

        for index, file in enumerate(filesList):
            self.__fileTree.insert("",ttk.END, values=(index+1,file))

    def __clearAllFiles(self):
        # Delete all files in the table
        for i in self.__fileTree.get_children():
            self.__fileTree.delete(i)

        self.__dataHandler.setFilesPathList([])

    def __deleteFile(self):
        # Delete user selected files
        selections = self.__fileTree.selection()        
        for item in selections:
            self.__fileTree.delete(item)

        newList = []
        i = 0
        for row in self.__fileTree.get_children():
            for filePath in self.__fileTree.item(row)["values"]:
                if i % 2 == 1:
                    newList.append(filePath)
                i+=1

        self.__dataHandler.setFilesPathList(newList)
        self.__displayFiles()

    def __showPage(self, frame):
        # Used to change pages
        if frame == self.__viewPage:
            self.geometry("1500x500+100+250")
            self.__dataHandler.setSchedules()
            self.__displayHandler = DisplayHandler(self.__dataHandler.getSchedules(), self.__dataHandler.getRangeList())
            self.__ascButton.invoke()
            self.__updateOptions()
            self.__displaySchedules()
        else:
            self.geometry("750x500+500+250")

        self.__currentFrame = frame
        frame.tkraise()

    def __displaySchedules(self):
        # Display sorted schedules in the table
        for i in self.__scheduleViewer.get_children():
                self.__scheduleViewer.delete(i)
        dayString = ["","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        try:
            for index, schedule in enumerate(self.__displayHandler.getResult()):
                li = schedule.data.getAll()
                li[7] = dayString[li[7]]
                li.insert(0,index+1)
                self.__scheduleViewer.insert("", tk.END, value=li)
        except TypeError:
            if self.__currentFrame is not self.__loadPage:
                mb.show_warning("No schedules to disply.", "No Files Warning")
            for i in self.__scheduleViewer.get_children():
                self.__scheduleViewer.delete(i)

    def __exportExcel(self):
        if self.__displayHandler.getResult() == None:
            mb.show_error("No schedules to export.\nPlease try again", "No schedule warning")
        else:
            try:
                path = fd.askdirectory()
                if path != "": 
                    self.__exportHandler.setResult(self.__displayHandler.exportResult(),path)
                    self.__exportHandler.exportExcel()
                    mb.show_info("Export successfully done.", "Export Success")
            except:
                mb.show_error("Export failed!\nPlease try again.", "Export Failure")

    def __exportPDF(self):
        if self.__displayHandler.getResult() == None:
            mb.show_error("No schedules to export.\nPlease try again", "No schedule warning")
        else:
            try:
                path = fd.askdirectory()
                if path != "":
                    self.__exportHandler.setResult(self.__displayHandler.exportResult(),path)
                    self.__exportHandler.exportPDF()
                    mb.show_info("Export successfully done.", "Export Success")
            except:
                mb.show_error("Export failed!\nPlease try again.", "Export Failure")

    def __insertionSort(self,array) -> list:
        newList = []
        for i in array: newList.append(i)

        for i in range(1,len(newList)):
            value = newList[i]
            pointer = i
            while pointer > 0 and newList[pointer-1] > value:
                newList[pointer] = newList[pointer-1]
                pointer -=1
            newList[pointer] = value

        return newList



# ===== Main =====
if __name__ == "__main__":
    program = GUI()