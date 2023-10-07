from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color

course = ["SET Introduction to Programming (IP)", "SET Introduction to Programming (IP)", "SET Introduction to Programming (IP)"]
courseNum = len(course)

courseAbbr = ["IP", "DCNG", "ANBCE"]
courseAbbrNum = len(courseAbbr)

lecturer = ["MNIA ASDASDASD", "asdasdas sdfasf", "asd asdads"]
lecturerNum = len(lecturer)

courseLocation = ["abcde & asdfads", "asdasd", "sadaas"]

wb = Workbook()
ws = wb.active

#Setting the header of the schedule

# Width Setting
ws.column_dimensions['A'].width = 5

col = 2
while col < 20:
    i = get_column_letter(col)
    ws.column_dimensions[i].width = 15
    col += 1

headFont = Font(bold = True, size=14)
headUnderFont = Font(bold = True, underline = "single", size=14)

ws['B2'].font = headFont
ws['B2'] = "PSB Academy"

ws['B3'].font = Font(size=12)
ws['B3'] = " Marina Square"
ws.merge_cells('B3:C3')

ws['B4'].font = Font(size=12)
ws['B4'] = " 6 Raffles Blvd, #03-200"
ws.merge_cells('B4:C4')

ws['B5'].font = Font(size=12)
ws['B5'] = " Singapore 039594"
ws.merge_cells('B5:C5')

ws['B8'].font = headFont
ws['B8'] = "Program Title"

for i, n in enumerate(course):
    ws.merge_cells('C' + str(i+8) +":"+'H' + str(i+8))
    ws['C' + str(i+8)].font = Font(size=14)
    ws['C' + str(i+8)] = n

ws['B' + str(courseNum+10)].font = headUnderFont
ws['B' + str(courseNum+10)] = "Module Abbr."

for i, n in enumerate(courseAbbr):
    ws.merge_cells('B' + str(i+courseNum+11) +":"+'C' + str(i+courseNum+11))
    ws['B' + str(i+courseNum+11)].font = Font(size=12)
    ws['B' + str(i+courseNum+11)] = " "+n

ws['D' + str(courseNum+10)].font = headUnderFont
ws['D' + str(courseNum+10)] = "Lecturer"

for i, n in enumerate(lecturer):
    ws.merge_cells('D' + str(i+courseNum+11) +":"+'E' + str(i+courseNum+11))
    ws['D' + str(i+courseNum+11)].font = Font(size=12)
    ws['D' + str(i+courseNum+11)] = " "+n

ws['F' + str(courseNum+10)].font = headUnderFont
ws['F' + str(courseNum+10)] = "Class Type"

for i, n in enumerate(courseLocation):
    ws.merge_cells('F' + str(i+courseNum+11) +":"+'I' + str(i+courseNum+11))
    ws['F' + str(i+courseNum+11)].font = Font(size=12)
    ws['F' + str(i+courseNum+11)] = " "+n



ws.sheet_view.showGridLines = False
ws.sheet_view.showRowColHeaders = False

wb.save("Hello2.xlsx")