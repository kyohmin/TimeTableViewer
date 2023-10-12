from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
import os, csv
import datetime
import colorsys
import random

class Node:
    def __init__(self, data):
        self.data = data
        self.next = None

    def __iter__(self):
        currentNode = self
        while currentNode:
            yield currentNode
            currentNode = currentNode.next

class LinkedList:
    def __init__(self):
        self.head = None # Starting point of the linked list
        self.tail = None # Tail is added for O(1) for appending
    
    def __iter__(self):
        currentNode = self.head
        while currentNode:
            yield currentNode
            currentNode = currentNode.next

    def __contains__(self, node):
        currentNode = self.head
        while currentNode:
            if currentNode.data == node:
                return True
            currentNode = currentNode.next

    def append(self, data):
        newNode = Node(data)
        if self.head == None:
            self.head = newNode
            self.tail = newNode
        
        else:
            self.tail.next = newNode
            self.tail= newNode

class Schedule():
    def __init__(self, module:str, moduleCode:str, cohort:str, course:str, fullPart:str,
                session:str, activityDate:str, scheduledDay:str,startTime:str,endTime:str,
                duration:str,location:str,size:str,lecturer:str,zone:str):
        self.__module = module
        self.__moduleCode = moduleCode
        self.__cohort = cohort
        self.__course = course
        self.__fullPart = fullPart
        self.__session = session
        self.__activityDate = activityDate
        self.__scheduledDay = scheduledDay
        self.__startTime = startTime
        self.__endTime = endTime
        self.__duration = duration
        self.__location = location
        self.__size = size
        self.__lecturer = lecturer
        self.__zone = zone
    
    def get(self, key): # Used for specific value
        value = {
            "module" : self.__module, "moduleCode" : self.__moduleCode, "cohort" : self.__cohort,
            "course" : self.__course, "fullPart" : self.__fullPart, "session" : self.__session,
            "activityDate" : self.__activityDate, "scheduledDay" : self.__scheduledDay,
            "startTime" : self.__startTime, "endTime" : self.__endTime, "duration" : self.__duration,
            "location" : self.__location, "size" : self.__size, "lecturer" : self.__lecturer, "zone" : self.__zone
        }[key]
        
        return value
    
    def getAll(self) -> list: # Used for all values
        li = [
            self.__module, self.__moduleCode, self.__cohort, self.__course, self.__fullPart,
            self.__session, self.__activityDate, self.__scheduledDay, self.__startTime,
            self.__endTime, self.__duration, self.__location, self.__size, self.__lecturer, self.__zone
        ]

        return  li
        


class DataHandler:
    def __init__(self):
        self.__schedules = LinkedList()
        self.__filesPathList = ['/Users/khms/CODE/ALGO/dataset/dataset 2020/220_FT ISOG.csv', '/Users/khms/CODE/ALGO/dataset/dataset 2020/220_FT IP.csv', '/Users/khms/CODE/ALGO/dataset/dataset 2020/220_FT DCNG.csv', '/Users/khms/CODE/ALGO/dataset/dataset 2020/120_FT DDMG.csv', '/Users/khms/CODE/ALGO/dataset/dataset 2020/120_FT DM.csv', '/Users/khms/CODE/ALGO/dataset/dataset 2020/120_FT ICOS.csv']
        self.__rangeList = [] # For initializing hash table. Contains below values in int

        # To hand over the range number to DisplayHandler for intialization of Hash Table
        self.__moduleRange = set()
        self.__moduleCodeRange = set()
        self.__cohortRange = set()
        self.__courseRange = set()
        self.__fullPartRange = set()
        self.__sessionRange = set()
        self.__activityDateRange = set()
        self.__scheduledDayRange = set()
        self.__startTimeRange = set()
        self.__endTimeRange = set()
        self.__durationRange = set()
        self.__locationRange = set()
        self.__sizeRange = set()
        self.__lecturerRange = set()
        self.__zoneRange = set()

    def setSchedules(self):
        self.__schedules.head = None  # Reset the head before storing the schedules
        self.__resetRange() # Reset the range before storing the schedules

        # Read CSV files from the path
        for fileDirectory in self.__filesPathList:
            with open(fileDirectory, "r") as csvfile:
                csvData = csv.reader(csvfile, delimiter=",")
                for row in csvData:
                    if row[0] == '':
                        continue

                    # Extract certain values from the raw data
                    module, moduleCode, cohort, course, fullPart, session = self.__extractData(row[1], row[2])

                    # Set range
                    self.__setRange(
                        module, moduleCode, cohort, course, fullPart, session, self.__dateInput(row[3]),
                        self.__dateInput(row[3]).isoweekday(), self.__timeInput(row[5]),
                        self.__timeInput(row[6]),row[7].lstrip("0"), row[8], int(row[9]), row[10], row[11]
                    )
                    
                    # Store schedule class into linked list
                    self.__schedules.append(
                        Schedule(
                            module, moduleCode, cohort, course, fullPart, session, self.__dateInput(row[3]),
                            self.__dateInput(row[3]).isoweekday(), self.__timeInput(row[5]), self.__timeInput(row[6]),
                            row[7].lstrip("0"), row[8], int(row[9]), row[10], row[11]
                        )
                    )

        # Storing size of range for optimal hashtable size
        self.__rangeList = [len(self.__moduleRange), len(self.__moduleCodeRange), len(self.__cohortRange), len(self.__courseRange), len(self.__fullPartRange), 
                            len(self.__sessionRange), len(self.__activityDateRange), len(self.__scheduledDayRange), len(self.__startTimeRange), len(self.__endTimeRange), 
                            len(self.__durationRange), len(self.__locationRange), len(self.__sizeRange), len(self.__lecturerRange), len(self.__zoneRange)]

    def getSchedules(self) -> Node:
        return self.__mergeSort(self.__schedules.head,"activityDate")
    
    def getRangeList(self) -> list:
        return self.__rangeList

    def setFilesPathList(self, filesList:list):
        self.__filesPathList = filesList

    def getFilesPathList(self) -> list:
        return self.__filesPathList

    def setRange(self, schedules):
        # Used externally for finding the range of each values
        self.__resetRange()
        try:
            for i in schedules:
                module, moduleCode, cohort, course, fullPart, session, activityDate, scheduleDay, startTime, endTime, duration, location, size, lecturer, zone = i.data.getAll()
                self.__setRange(module, moduleCode, cohort, course, fullPart, session, activityDate, scheduleDay, startTime, endTime, duration, location, size, lecturer, zone)
        except:
            pass

    def getRange(self, category) -> set: # Used for available options
        value = {
        "module": self.__moduleRange,
        "moduleCode": self.__moduleCodeRange,
        "cohort": self.__cohortRange,
        "course": self.__courseRange,
        "fullPart": self.__fullPartRange,
        "session": self.__sessionRange,
        "activityDate": self.__activityDateRange,
        "scheduledDay": self.__scheduledDayRange,
        "startTime": self.__startTimeRange,
        "endTime": self.__endTimeRange,
        "duration": self.__durationRange,
        "location": self.__locationRange,
        "size": self.__sizeRange,
        "lecturer": self.__lecturerRange,
        "zone": self.__zoneRange,
        }[category]

        return value

    def __dateInput(self, strVal):
        day, month, year = strVal.split("/")
        return datetime.date(int(year),int(month),int(day))

    def __timeInput(self, strVal):
        hour, minute,second = strVal.split(":")
        return datetime.time(int(hour),int(minute),int(second))

    def __setRange(self, module:str, moduleCode:str, cohort:str, course:str, fullPart:str,
                   session:str, activityDate:str, scheduleDay:str, startTime:str, endTime:str,
                   duration:str, location:str, size:str, lecturer:str, zone:str):
        # Internally used for initialization of hash table
        self.__moduleRange.add(module)
        self.__moduleCodeRange.add(moduleCode)
        self.__cohortRange.add(cohort)
        self.__courseRange.add(course)
        self.__fullPartRange.add(fullPart)
        self.__sessionRange.add(session)
        self.__activityDateRange.add(activityDate)
        self.__scheduledDayRange.add(scheduleDay)
        self.__startTimeRange.add(startTime)
        self.__endTimeRange.add(endTime)
        self.__durationRange.add(duration)
        self.__locationRange.add(location)
        self.__sizeRange.add(size)
        self.__lecturerRange.add(lecturer)
        self.__zoneRange.add(zone)

    def __extractData(self,name:str, description:str) -> list:
        course,cohort,fullPart,moduleCode,session = name.split("_")
        cohort = fullPart + "_" + cohort
        module = description[4:].split(" (")[0]
        
        return [module, moduleCode, cohort, course, fullPart, session]
    
    def __resetRange(self):
        self.__moduleRange.clear()
        self.__moduleCodeRange.clear()
        self.__cohortRange.clear()
        self.__courseRange.clear()
        self.__fullPartRange.clear()
        self.__sessionRange.clear()
        self.__activityDateRange.clear()
        self.__scheduledDayRange.clear()
        self.__startTimeRange.clear()
        self.__endTimeRange.clear()
        self.__durationRange.clear()
        self.__locationRange.clear()
        self.__sizeRange.clear()
        self.__lecturerRange.clear()
        self.__zoneRange.clear() 

    def __mergeSort(self, head:Node, category:str) -> Node:
        if head is None or head.next is None:
            return head

        mid = self.__split(head)
        left = head
        right = mid.next
        mid.next = None 

        left = self.__mergeSort(left, category)
        right = self.__mergeSort(right, category)

        sortedHead = self.__merge(left, right, category)

        return sortedHead

    def __split(self, head:Node):
        half = head
        end = head

        while end.next is not None and end.next.next is not None:
            half = half.next
            end = end.next.next

        return half

    def __merge(self, left:Node, right:Node, category:str) -> Node: 
        startNode = Node(None)
        currentNode = startNode

        # Loop through to sort between two linked lists
        while left is not None and right is not None:
            if left.data.get(category) < right.data.get(category):
                currentNode.next = left
                left = left.next
            else:
                currentNode.next = right
                right = right.next

            currentNode = currentNode.next

        # After One side is done sorting, add the rest values
        if left is not None:
            currentNode.next = left
        elif right is not None:
            currentNode.next = right

        return startNode.next  # Because first Node was used to create rest of the nodes


class ResultHandler:
    def __init__(self):
        self.__reuslt = LinkedList()
        self.__startingRow = 12
        self.__weekCnt = 1
        self.__minDate = datetime.date(3000,1,1)
        self.__maxDate = datetime.date(1000,1,1)
        self.__groupedModules = {}
        self.__path = ""

    def setResult(self, result):
        self.__reuslt = result

    def __resetData(self):
        self.__reuslt = LinkedList()
        self.__startingRow = 12
        self.__weekCnt = 1
        self.__minDate = datetime.date(3000,1,1)
        self.__maxDate = datetime.date(1000,1,1)
        self.__groupedModules = {}
        self.__path = ""

    def exportExcel(self):
        # Reset all data before exporting
        self.__resetData()

        # Border width predefine
        self.__thick = Side(border_style="medium")
        self.__thin = Side(border_style="thin",color="808080")
        self.__getHeaderData()

        wb = Workbook()
        ws = wb.active

        # First Row with small width
        ws.column_dimensions['A'].width = 5

        # Default Row width
        col = 2
        while col < 10:
            i = get_column_letter(col)
            ws.column_dimensions[i].width = 30
            col += 1

        # PSB logo
        ws.merge_cells('B4:B7')
        ws['B4'].font = Font(size=76,color="902108",name='Times New Roman', bold=True)
        ws['B4'].alignment = Alignment(horizontal="center",vertical="center")
        ws['B4'] = "PSB"

        # ACADEMY logo
        ws.row_dimensions[8].height = 20
        ws.row_dimensions[9].height = 20
        ws.merge_cells('B8:B9')
        ws['B8'].font = Font(size=36,color="7F7F7F",name='Calibri', bold=True)
        ws['B8'].alignment = Alignment(horizontal="center",vertical="center")
        ws['B8'] = "Academy"

        # Grouped Module list header
        ws.merge_cells('D4:E4')
        ws['D4'].alignment = Alignment(horizontal="left",vertical="center")
        ws['D4'] = "Module"
        ws['F4'] = "Cohort"
        ws['G4'] = "Lecturer"
        ws['H4'] = "Full-time / Part-time"
        ws['I4'] = "Date Start - Date End"

        ws['D4'].border = Border(left=self.__thick,top=self.__thick,bottom=self.__thick)
        ws['E4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['F4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['G4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['H4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thin)
        ws['I4'].border = Border(top=self.__thick,bottom=self.__thick,right=self.__thick)

        ws['D4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['F4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['G4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['H4'].fill = PatternFill("solid", start_color="AACDFF")
        ws['I4'].fill = PatternFill("solid", start_color="AACDFF")

        # Grouped Module Data Insert
        i = 5
        for _, value in self.__groupedModules.items():
            module, cohort, lecturer, fullPart = value[0], value[1], value[2], value[3]
            ws.merge_cells('D'+str(i)+':'+"E"+str(i))
            ws["D" + str(i)] = module
            ws["F" + str(i)] = cohort
            ws["G" + str(i)] = lecturer
            ws["H" + str(i)] = fullPart
            ws["D" + str(i)].fill = PatternFill("solid", start_color=value[4])

            ws["D" + str(i)].border = Border(left=self.__thick,bottom=self.__thin)
            ws["E" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["F" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["G" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["H" + str(i)].border = Border(bottom=self.__thin,right=self.__thin)
            ws["I" + str(i)].border = Border(bottom=self.__thin,right=self.__thick)
            if len(self.__groupedModules.items())-1 == i-5:
                ws["D" + str(i)].border = Border(left=self.__thick,bottom=self.__thick)
                ws["E" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["F" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["G" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["H" + str(i)].border = Border(bottom=self.__thick,right=self.__thin)
                ws["I" + str(i)].border = Border(bottom=self.__thick,right=self.__thick)
            ws.row_dimensions[i].height = 20
            i+=1

        # Define the Starting row
        self.__startingRow = 12
        if len(self.__groupedModules.items()) > 5:
            self.__startingRow += len(self.__groupedModules.items())-4

        # Build Main Schedule table
        self.currentDate = self.__minDate
        self.weekStartMarked = False
        self.__createHeading(ws)
        self.__createRow(ws)
        moduleRow = self.__startingRow
        createdRow = 1
        stream = 1
        prevDate = self.__minDate - datetime.timedelta(days=1)

        # Insert schedules
        for schedule in self.__reuslt:
            date = schedule.data.get("activityDate")
            # if new heading has to be created
            while date >= self.currentDate:
                date = schedule.data.get("activityDate")
                self.__createHeading(ws)
                self.__createRow(ws)
                moduleRow = self.__startingRow
                createdRow = 1
                stream = 1

            # if the schedule is the first to go in the date
            if date != prevDate:
                stream = 1
                insertingCell = get_column_letter(date.isoweekday()+2) + str(moduleRow - 6)
                ws[insertingCell] = schedule.data.get("moduleCode") + " (" + schedule.data.get("fullPart")+")"
                ws[insertingCell].fill = PatternFill("solid", start_color=self.__groupedModules[f"{schedule.data.get('module')}{schedule.data.get('cohort')}{schedule.data.get('lecturer')}{schedule.data.get('fullPart')}"][4])
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 5)] = str(schedule.data.get("startTime"))[:-3] + " ~ " + str(schedule.data.get("endTime"))[:-3]
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 4)] = self.__characterLimit(schedule.data.get("lecturer"))
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 3)] = schedule.data.get("location")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 2)] = schedule.data.get("size")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 1)] = schedule.data.get("zone")
            
            # If the schedule has to go to second ~ # rows
            else:
                stream += 1
                # If I need to create a new row for new data
                if createdRow < stream:
                    self.__createRow(ws)
                    createdRow += 1
                # Insert schedule
                insertingCell = get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6))
                ws[insertingCell] = schedule.data.get("moduleCode") + " (" + schedule.data.get("fullPart")+")"
                ws[insertingCell].fill = PatternFill("solid", start_color=self.__groupedModules[f"{schedule.data.get('module')}{schedule.data.get('cohort')}{schedule.data.get('lecturer')}{schedule.data.get('fullPart')}"][4])
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+1)] = str(schedule.data.get("startTime"))[:-3] + " ~ " + str(schedule.data.get("endTime"))[:-3]
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+2)] = self.__characterLimit(schedule.data.get("lecturer"))
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+3)] = schedule.data.get("location")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+4)] = schedule.data.get("size")
                ws[get_column_letter(date.isoweekday()+2) + str(moduleRow - 12 + (stream*6)+5)] = schedule.data.get("zone")
            prevDate = date


        # Remove the frid lines and headers in Excel for clean view
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = False

        wb.save(self.__path+"Excel_Timetable.xlsx")

    def __getHeaderData(self): # Used for finding the grouped schedules
        groupingSchedule = {}
        for i in self.__reuslt:
            strVal = f"{i.data.get('module')}{i.data.get('cohort')}{i.data.get('lecturer')}{i.data.get('fullPart')}"
            groupingSchedule.update({strVal:[i.data.get('module')+' '+i.data.get('moduleCode'),i.data.get('cohort'),i.data.get('lecturer'),i.data.get('fullPart'),self.__generateRandomColor()]})
            if i.data.get("activityDate") < self.__minDate: self.__minDate = i.data.get("activityDate")
            if i.data.get("activityDate") > self.__maxDate: self.__maxDate = i.data.get("activityDate")

        self.__groupedModules = groupingSchedule
    
    def __createHeading(self, ws): # Used for creating heading with days and dates
        # 
        headingFont = Font(size=14, bold=True, underline="single")
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for i in range(8):
            currentColumnLetter = get_column_letter(i+2)
            currentCell = currentColumnLetter + str(self.__startingRow)
            bottomCell = currentColumnLetter + str(self.__startingRow+1)
            
            # For the Week columns
            if currentColumnLetter == "B":
                ws.merge_cells('B' + str(self.__startingRow)+':'+'B'+ str(self.__startingRow+1))
                ws[currentCell] = "Week" + str(self.__weekCnt)
                ws[currentCell].border = Border(left=self.__thick, top=self.__thick,right=self.__thick)
                ws[bottomCell].border = Border(left=self.__thick, bottom=self.__thick, right=self.__thick)

            # For the Days and date columns
            else:
                currentCell = currentColumnLetter + str(self.__startingRow)
                ws[currentCell] = days[i-1]
                ws[currentCell].border = Border(top=self.__thick,right=self.__thin)
                ws[bottomCell].border = Border(bottom=self.__thick, right=self.__thin)
                if currentColumnLetter == "I":
                    ws[currentCell].border = Border(top=self.__thick,right=self.__thick)
                    ws[bottomCell].border = Border(bottom=self.__thick,right=self.__thick)
                

            # Font and color on all cells
            ws[currentCell].font = headingFont
            ws[currentCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[bottomCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
            ws[bottomCell].alignment = Alignment(horizontal="center",vertical="center")
        
        # Entering Date
        
        for i in range(7):
            currentColumnLetter = get_column_letter(i+3)
            currentCell = currentColumnLetter + str(self.__startingRow+1)
            if self.__weekCnt == 1 and not self.weekStartMarked:
                if self.currentDate.isoweekday() == i+1:
                    ws[currentCell] = self.currentDate
                    self.weekStartMarked = True
                
            # If start has marked
            else:
                ws[currentCell] = self.currentDate

            if self.weekStartMarked == True:
                self.currentDate += datetime.timedelta(days=1)

        self.__weekCnt += 1
        self.__startingRow += 2

    def __createRow(self,ws): # Used for creating empty schedules row frame
        detailList = ["Module", "Time", "Lecturer", "Location", "Size", "Zone"]
        ws.row_dimensions[self.__startingRow].height = 70
        # Left Column border and data setting
        for i in range(6):
            currentCell = "B" + str(i + self.__startingRow)
            ws[currentCell] = detailList[i]
            ws[currentCell].fill = PatternFill("solid", start_color="AACDFF")
            ws[currentCell].font = Font(bold = True, size=14)
            ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
            ws[currentCell].border = Border(bottom=self.__thin,right=self.__thick,left=self.__thick)

        ws["B" + str(5 + self.__startingRow)].border = Border(bottom=self.__thick,right=self.__thick,left=self.__thick)

        # Setting the font and border for the actual values
        for i in range(7):
            for j in range(6):
                currentColumnLetter = get_column_letter(i+3)
                currentCell = currentColumnLetter + str(j+self.__startingRow)
                ws[currentCell].font = Font(size=14)
                ws[currentCell].border = Border(bottom=self.__thin,right=self.__thin)
                ws[currentCell].alignment = Alignment(horizontal="center",vertical="center")
                if j == 5:
                    ws[currentCell].border = Border(bottom=self.__thick,right=self.__thin)
                if i == 6:
                    ws[currentCell].border = Border(bottom=self.__thin,right=self.__thick)
                if j == 5 and i == 6:
                    ws[currentCell].border = Border(bottom=self.__thick,right=self.__thick)

        self.__startingRow += 6

    def __characterLimit(self, strVal) -> str:
        # If the character is too big, only display first and last name
        if len(strVal) > 25:
            return strVal.split()[0] + " "+ strVal.split()[-1]
        else:
            return strVal
        
    def __generateRandomColor(self) -> str:
        # Generate random color for grouping modules
        h,s,l = random.random()*360, random.random(), 0.9
        r,g,b = [int(256*i) for i in colorsys.hls_to_rgb(h,l,s)]
        hexVal = '{:02x}{:02x}{:02x}'.format(r, g, b)

        return hexVal
    
    

dataHandler = DataHandler()
dataHandler.setSchedules()
resultHandler = ResultHandler()
resultHandler.setResult(dataHandler.getSchedules())
resultHandler.exportExcel()